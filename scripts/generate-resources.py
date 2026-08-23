from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "assets" / "downloads"
RESOURCES = json.loads((ROOT / "data" / "resources.json").read_text(encoding="utf-8"))["resources"]
NAVY = "185A77"
TEAL = "2F83A6"
GOLD = "D6A02B"
MIST = "F4F8FB"
LIGHT = "E6F0F5"
INPUT = "FFF2CC"
THIN = Side(style="thin", color="B7CFDA")

RESOURCE_ROWS = {
    "iso-9001-hazirlik-kontrol-listesi": [
        ("Bağlam", "İç/dış konular ve ilgili taraf beklentileri güncel mi?", "Bağlam analizi, paydaş listesi"),
        ("Süreçler", "Süreç sahipleri, girdiler, çıktılar ve etkileşimler tanımlı mı?", "Süreç haritası, görev matrisi"),
        ("Riskler", "Risk ve fırsatlar aksiyon, sorumlu ve tarihle izleniyor mu?", "Risk kaydı, aksiyon planı"),
        ("Hedefler", "Kalite hedefleri ölçülebilir ve süreçlerle ilişkili mi?", "Hedef/KPI tablosu"),
        ("Kaynaklar", "Yetkinlik, altyapı ve çalışma ortamı gereksinimleri kontrol ediliyor mu?", "Eğitim, bakım, kalibrasyon kaydı"),
        ("Operasyon", "Müşteri şartı, değişiklik ve dış sağlayıcı kontrolleri uygulanıyor mu?", "Sipariş, değişiklik, tedarikçi kayıtları"),
        ("Performans", "Şikâyet, uygunsuzluk ve performans eğilimleri düzenli değerlendiriliyor mu?", "Gösterge, DÖF, analiz"),
        ("İç tetkik", "Tetkik kapsamı, tarafsızlık, bulgu ve kapanış kanıtı tamam mı?", "Program, rapor, kapanış kanıtı"),
        ("YGG", "Yönetim gözden geçirmesi gerekli girdilerle karar üretiyor mu?", "Gündem, tutanak, karar takibi"),
    ],
    "iso-45001-saha-denetim-kontrol-listesi": [
        ("Makine", "Koruyucular, acil durdurma ve tehlikeli bölge erişimi kontrol edildi mi?", "Fotoğraf, ekipman kontrol kaydı"),
        ("LOTO", "Bakım/temizlik/ayar işlerinde tüm enerji kaynakları tanımlı mı?", "LOTO kartı, izolasyon doğrulaması"),
        ("Kimyasal", "SDS, etiket, depolama ve dökülme ekipmanı güncel mi?", "Kimyasal envanter, saha kontrolü"),
        ("Yüklenici", "Yeterlilik, saha kabulü, izin ve gözetim adımları uygulanıyor mu?", "Kabul formu, izin, gözlem"),
        ("Trafik", "Araç-yaya ayrımı, kör noktalar ve hız kuralları görünür mü?", "Saha planı, gözlem kaydı"),
        ("Yüksekte çalışma", "Erişim, ankraj, ekipman ve kurtarma düzeni doğrulandı mı?", "İzin, ekipman kontrolü, plan"),
        ("Acil durum", "Kaçış yolları, ekipman, görevler ve iletişim erişilebilir mi?", "Plan, kontrol, tatbikat kaydı"),
        ("Aksiyon", "Bulgu sahibine, tarihe ve kapanış kanıtına bağlandı mı?", "Aksiyon kaydı, doğrulama"),
    ],
    "30-gunluk-denetim-hazirlik-plani": [
        ("1. hafta", "Kapsam, lokasyon, standart ve denetim tarihini doğrulayın", "Onaylı kapsam"),
        ("1. hafta", "Süreç sahipleriyle kanıt listesi ve sorumluları netleştirin", "Sorumluluk matrisi"),
        ("1. hafta", "Ön boşluk taraması ve kritik açık önceliklendirmesi yapın", "Açık aksiyon planı"),
        ("2. hafta", "Zorunlu kayıtların güncellik ve erişilebilirliğini doğrulayın", "Kanıt matrisi"),
        ("2. hafta", "Saha uygulaması ve doküman eşleşmesini örneklemle test edin", "Saha gözlem kaydı"),
        ("2. hafta", "Geciken DÖF ve risk aksiyonlarını sahipleriyle kapatın", "Kapanış kanıtları"),
        ("3. hafta", "İç tetkik veya prova denetimi yürütün", "Tetkik raporu"),
        ("3. hafta", "Bulgular için kök neden, düzeltme ve etkinlik planı oluşturun", "DÖF kayıtları"),
        ("4. hafta", "YGG girdilerini ve karar takibini tamamlayın", "YGG tutanağı"),
        ("4. hafta", "Son kanıt indeksi, görüşme planı ve lojistiği doğrulayın", "Denetim dosyası"),
    ],
    "denetim-kanit-matrisi": [
        ("Madde", "Gerekliliği süreç ve sorumluya bağlayın", "Standart maddesi / müşteri şartı"),
        ("Süreç", "Gerekliliğin uygulandığı gerçek operasyonu yazın", "Süreç haritası"),
        ("Kanıt", "Belge adı yerine erişilebilir gerçek kaydı tanımlayın", "Kayıt bağlantısı / dosya yolu"),
        ("Güncellik", "Kanıtın dönem, revizyon ve geçerlilik tarihini doğrulayın", "Revizyon/geçerlilik"),
        ("Örneklem", "Farklı vardiya, ürün, lokasyon veya çalışan örnekleyin", "Örneklem listesi"),
        ("Açık", "Eksik kanıtı sahibine, tarihe ve doğrulama yöntemine bağlayın", "Aksiyon kaydı"),
    ],
    "acil-durum-tatbikat-raporu": [
        ("Senaryo", "Olay türü, başlangıç noktası ve varsayımlar tanımlı mı?", "Onaylı senaryo"),
        ("Bildirim", "Alarm ve iç/dış iletişim zamanları kaydedildi mi?", "Kronoloji"),
        ("Tahliye", "Kaçış, sayım ve özel destek ihtiyacı gözlendi mi?", "Gözlem formu"),
        ("Ekipler", "Görev, ekipman ve iletişim performansı değerlendirildi mi?", "Ekip değerlendirmesi"),
        ("Süre", "Kritik aşamaların başlangıç/bitiş zamanı ölçüldü mü?", "Zaman çizelgesi"),
        ("Aksiyon", "Bulgu, sorumlu, termin ve doğrulama yöntemi belirlendi mi?", "Aksiyon planı"),
    ],
    "iso-50001-seu-enpi-calisma-kitabi": [
        ("Sınır", "Enerji yönetim sistemi sınırı ve veri kaynakları tanımlı mı?", "Kapsam ve sayaç listesi"),
        ("SEU", "Tüketim ve iyileştirme potansiyeline göre önemli kullanımlar seçildi mi?", "SEU değerlendirmesi"),
        ("Değişken", "Üretim, derece-gün, çalışma saati gibi etkileyen değişkenler belirlendi mi?", "Veri sözlüğü"),
        ("Baz hat", "Temsili dönem ve normalizasyon yöntemi gerekçelendirildi mi?", "Baz hat modeli"),
        ("EnPI", "Gösterge birimi, formülü, sorumlusu ve sıklığı açık mı?", "EnPI kartı"),
        ("Doğrulama", "Proje öncesi/sonrası veri ve dış etkenler birlikte değerlendiriliyor mu?", "Tasarruf doğrulama kaydı"),
    ],
    "tekstil-izlenebilirlik-kutle-denge-kontrol-listesi": [
        ("Tedarikçi", "Belge kapsamı, ürün, tesis ve geçerlilik doğrulandı mı?", "Onaylı tedarikçi kanıtı"),
        ("Hammadde", "Giriş miktarı, lot, iddia ve depo kaydı aynı kodla izleniyor mu?", "Mal kabul ve stok kaydı"),
        ("Üretim", "Sipariş, reçete, üretim kaybı ve yeniden işleme kaydediliyor mu?", "Üretim ve fire kaydı"),
        ("Kütle dengesi", "Girdi, fire, stok ve çıktı seçilen dönem için uzlaşıyor mu?", "Kütle denge tablosu"),
        ("Sevkiyat", "Ürün iddiası, miktar ve satış/sevkiyat kaydı eşleşiyor mu?", "Fatura ve sevkiyat kaydı"),
        ("Örneklem", "Rastgele sipariş geriye ve ileriye doğru izlenebiliyor mu?", "İzlenebilirlik prova kaydı"),
    ],
    "iso-27001-risk-varlik-envanteri-kontrol-listesi": [
        ("Varlıklar", "Bilgi, uygulama, cihaz, hizmet ve tesis varlıkları envanterde mi?", "Varlık envanteri, sahiplik kaydı"),
        ("Sınıflandırma", "Varlıkların önem, gizlilik ve erişilebilirlik seviyesi tanımlı mı?", "Sınıflandırma kuralı, kayıt"),
        ("Risk", "Tehdit, zafiyet, etki ve olasılık birlikte değerlendiriliyor mu?", "Risk değerlendirmesi"),
        ("Kontroller", "Risklere karşı teknik, fiziksel ve yönetsel kontroller atanmış mı?", "Kontrol planı, sorumlu matrisi"),
        ("Erişim", "Yetki talebi, onay, gözden geçirme ve iptal kayıtları tutuluyor mu?", "Erişim kayıtları"),
        ("Olay", "Bilgi güvenliği olayları sınıflandırılıyor ve öğrenimler aksiyona bağlanıyor mu?", "Olay kaydı, aksiyon planı"),
        ("Süreklilik", "Yedekleme, kurtarma ve kritik hizmet bağımlılıkları test ediliyor mu?", "Test raporu, kurtarma kaydı"),
        ("İyileştirme", "İç tetkik ve yönetim kararları risk kaydına geri besleniyor mu?", "Tetkik ve YGG kanıtı"),
    ],
    "iso-22301-is-etki-analizi-kontrol-listesi": [
        ("Faaliyet", "Ürün, hizmet ve kritik faaliyet sahipleri tanımlı mı?", "Faaliyet envanteri"),
        ("Etki", "Finansal, operasyonel, yasal ve itibar etkileri dönemlere göre değerlendirildi mi?", "Etki değerlendirmesi"),
        ("Tolerans", "Maksimum kabul edilebilir kesinti ve veri kaybı süresi belirlendi mi?", "MTPD/RPO kaydı"),
        ("Bağımlılık", "İnsan, teknoloji, tesis, tedarikçi ve veri bağımlılıkları haritalandı mı?", "Bağımlılık matrisi"),
        ("Öncelik", "Kurtarma sırası ve asgari kaynak ihtiyacı onaylandı mı?", "Önceliklendirme tablosu"),
        ("Strateji", "Alternatif çalışma, iletişim ve kaynak stratejileri tanımlı mı?", "Süreklilik stratejisi"),
        ("Test", "Planlar tatbikat veya senaryo ile doğrulanıp öğrenimler kaydediliyor mu?", "Tatbikat raporu"),
        ("Gözden geçirme", "Değişiklikler ve yönetim kararları analiz kayıtlarına işleniyor mu?", "Gözden geçirme tutanağı"),
    ],
    "iso-10002-sikayet-yonetimi-kontrol-listesi": [
        ("Kanal", "Şikâyetlerin tüm kanallardan alınacağı ve kaydedileceği tanımlı mı?", "Kanal ve kayıt prosedürü"),
        ("Sınıflandırma", "Konu, önem, müşteri ve sorumlu bilgisi tutarlı biçimde atanıyor mu?", "Şikâyet kayıtları"),
        ("İnceleme", "Tarafsız inceleme, kanıt toplama ve iletişim adımları izleniyor mu?", "İnceleme notları"),
        ("Yanıt", "Müşteriye verilen yanıt, tarih ve karar gerekçesiyle kayıtlı mı?", "Yanıt kaydı"),
        ("Kök neden", "Tekrarlayan şikâyetler kök neden ve düzeltici aksiyona bağlanıyor mu?", "Kök neden/DÖF kaydı"),
        ("Süre", "Yanıt ve kapanış süreleri hedeflerle karşılaştırılıyor mu?", "SLA ve gösterge tablosu"),
        ("Memnuniyet", "Şikâyet sonrası geri bildirim ve eğilimler değerlendiriliyor mu?", "Geri bildirim analizi"),
        ("YGG", "Şikâyet eğilimleri yönetim kararlarına ve iyileştirmeye taşınıyor mu?", "YGG girdisi"),
    ],
    "iso-22000-gida-guvenligi-on-hazirlik-listesi": [
        ("Kapsam", "Ürün, proses, tesis, vardiya ve dış kaynak kapsamı açık mı?", "Kapsam ve akış şeması"),
        ("Hijyen", "Temizlik, haşere, kişisel hijyen ve çevresel koşullar izleniyor mu?", "Hijyen ve temizlik kayıtları"),
        ("Tehlike", "Biyolojik, kimyasal, fiziksel ve alerjen tehlikeler analiz edildi mi?", "Tehlike analizi"),
        ("Kontrol noktası", "Kritik limit, izleme sıklığı ve sapma tepkisi tanımlı mı?", "KKN planı, ölçüm kaydı"),
        ("Tedarikçi", "Hammadde şartları, kabul kriterleri ve tedarikçi performansı izleniyor mu?", "Tedarikçi değerlendirmesi"),
        ("İzlenebilirlik", "Lot, giriş, üretim ve sevkiyat kayıtları geriye/ileriye izlenebiliyor mu?", "İzlenebilirlik provası"),
        ("Geri çağırma", "Geri çekme/çağırma senaryosu test edilmiş ve öğrenimler kaydedilmiş mi?", "Tatbikat raporu"),
        ("Doğrulama", "İç tetkik, analiz sonuçları ve yönetim gözden geçirmesi tamam mı?", "Tetkik ve YGG kanıtı"),
    ],
    "iso-14064-ghg-hesaplama-veri-matrisi": [
        ("Sınır", "Kontrol yaklaşımı, tesisler ve raporlama dönemi açıkça tanımlı mı?", "Organizasyonel sınır kaydı"),
        ("Kaynak", "Kapsam 1, 2 ve uygun Kapsam 3 kaynakları listelendi mi?", "Emisyon kaynak envanteri"),
        ("Aktivite verisi", "Yakıt, elektrik, soğutucu, seyahat ve lojistik verileri kaynağıyla tutuluyor mu?", "Fatura, sayaç, tüketim kaydı"),
        ("Faktör", "Emisyon faktörü kaynağı, birimi ve sürümü kaydedildi mi?", "Faktör referansı"),
        ("Hesap", "Birim dönüşümü, formül ve yuvarlama yöntemi tekrarlanabilir mi?", "Hesaplama çalışma kâğıdı"),
        ("Veri kalitesi", "Eksik, tahmini ve doğrulanmış veriler ayrı işaretleniyor mu?", "Veri kalite notu"),
        ("Belirsizlik", "Önemli varsayımlar ve belirsizlik kaynakları açıklanıyor mu?", "Varsayım/belirsizlik kaydı"),
        ("Rapor", "Sonuçlar, dönem karşılaştırması ve sınırlar yönetim tarafından gözden geçirildi mi?", "GHG raporu, onay kaydı"),
    ],
    "gida-denetim-hazirlik-kanit-matrisi": [
        ("Personel", "Görev yetkinliği, hijyen eğitimi ve sağlık kontrolleri güncel mi?", "Eğitim ve yetkinlik kayıtları"),
        ("Tesis", "Akış, zonlama, bakım ve altyapı koşulları kayıtlarla izleniyor mu?", "Tesis kontrolü, bakım kaydı"),
        ("Temizlik", "Plan, kimyasal, uygulama ve doğrulama kayıtları eksiksiz mi?", "Temizlik ve doğrulama kaydı"),
        ("Proses", "Kritik parametreler, sapmalar ve serbest bırakma kararları izleniyor mu?", "Proses izleme kaydı"),
        ("Tedarikçi", "Hammadde şartları, kabul kontrolleri ve uygunsuzluklar yönetiliyor mu?", "Tedarikçi ve kabul kaydı"),
        ("Ekipman", "Ölçüm cihazları, bakım ve kalibrasyon durumu görünür mü?", "Bakım/kalibrasyon kaydı"),
        ("İzlenebilirlik", "Numune sipariş geriye ve ileriye doğru takip edilebiliyor mu?", "İzlenebilirlik prova kaydı"),
        ("Aksiyon", "Bulgu, sorumlu, termin ve etkinlik doğrulaması tamam mı?", "Aksiyon ve kapanış kanıtı"),
    ],
    "sosyal-uygunluk-denetim-hazirlik-kontrol-listesi": [
        ("Çalışan kayıtları", "İşe giriş, görev, yaş ve özlük kayıtları kontrollü erişimle tutuluyor mu?", "Personel kayıtları"),
        ("Çalışma süresi", "Vardiya, puantaj, fazla çalışma ve izin kayıtları birbiriyle uyumlu mu?", "Puantaj ve vardiya kaydı"),
        ("Ücret", "Ücret, yan hak ve ödeme kayıtları çalışan sözleşmeleriyle eşleşiyor mu?", "Bordro ve ödeme kaydı"),
        ("İSG", "Riskler, eğitimler, olaylar ve acil durum kontrolleri uygulanıyor mu?", "İSG kayıtları"),
        ("Özgür irade", "İşe alım, disiplin, ayrımcılık ve taciz şikâyeti süreçleri tanımlı mı?", "Politika ve şikâyet kaydı"),
        ("Çalışan sesi", "Görüşme ve geri bildirim kanalları güvenli ve misillemesiz mi?", "Görüşme/geri bildirim özeti"),
        ("Tedarik zinciri", "Tedarikçi beklentileri, değerlendirme ve aksiyonlar izleniyor mu?", "Tedarikçi taahhüdü ve takip"),
        ("Kapanış", "Bulgu, sorumlu, termin ve doğrulama kanıtı kayıtlı mı?", "DÖF ve kapanış kaydı"),
    ],
    "30-60-90-egitim-etkinlik-raporu": [
        ("Eğitim", "Konu, hedef kitle, tarih, format ve eğitmen bilgisi tanımlı mı?", "Program ve katılım kaydı"),
        ("Ön ölçüm", "Başlangıç bilgi veya beceri seviyesi aynı yöntemle kaydedildi mi?", "Ön test / gözlem formu"),
        ("30 gün", "Katılımcı öğrendiği yöntemi gerçek görevinde uyguladı mı?", "Uygulama görevi ve kanıt"),
        ("60 gün", "Yönetici davranış, kontrol kalitesi ve ekip katılımını gözlemledi mi?", "Yönetici gözlem formu"),
        ("90 gün", "Son ölçüm ve saha bulguları ön ölçümle karşılaştırıldı mı?", "Son test / saha karşılaştırması"),
        ("Aksiyon", "Eksik kalan davranış veya süreç için sorumlu ve tarih belirlendi mi?", "Aksiyon planı"),
        ("Etkinlik", "Ölçüm sonucu bir sonraki eğitim veya danışmanlık ihtiyacına bağlandı mı?", "Etkinlik değerlendirmesi"),
        ("Onay", "Rapor, eğitim sahibi ve yönetici tarafından gözden geçirildi mi?", "Onay ve paylaşım kaydı"),
    ],
}


def register_fonts():
    regular = Path(r"C:\Windows\Fonts\arial.ttf")
    bold = Path(r"C:\Windows\Fonts\arialbd.ttf")
    if regular.exists() and bold.exists():
        pdfmetrics.registerFont(TTFont("GowayArial", str(regular)))
        pdfmetrics.registerFont(TTFont("GowayArialBold", str(bold)))
        return "GowayArial", "GowayArialBold"
    return "Helvetica", "Helvetica-Bold"


FONT, FONT_BOLD = register_fonts()


def pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="GTitle", parent=styles["Title"], fontName=FONT_BOLD, fontSize=23, leading=27, textColor=colors.HexColor("#" + NAVY), alignment=TA_LEFT, spaceAfter=10))
    styles.add(ParagraphStyle(name="GSub", parent=styles["Normal"], fontName=FONT, fontSize=10, leading=15, textColor=colors.HexColor("#" + TEAL), spaceAfter=12))
    styles.add(ParagraphStyle(name="GBody", parent=styles["BodyText"], fontName=FONT, fontSize=9.2, leading=13.2, textColor=colors.HexColor("#" + NAVY)))
    styles.add(ParagraphStyle(name="GSmall", parent=styles["BodyText"], fontName=FONT, fontSize=7.5, leading=10.5, textColor=colors.HexColor("#" + NAVY)))
    styles.add(ParagraphStyle(name="GHead", parent=styles["Heading2"], fontName=FONT_BOLD, fontSize=13, leading=16, textColor=colors.HexColor("#" + NAVY), spaceBefore=10, spaceAfter=7))
    return styles


def footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#" + TEAL))
    canvas.line(18 * mm, 13 * mm, A4[0] - 18 * mm, 13 * mm)
    canvas.setFont(FONT, 7)
    canvas.setFillColor(colors.HexColor("#" + NAVY))
    canvas.drawString(18 * mm, 8.5 * mm, "Goway Danışmanlık · Uygulama kaynağı")
    canvas.drawRightString(A4[0] - 18 * mm, 8.5 * mm, f"Sayfa {doc.page}")
    canvas.restoreState()


def make_pdf(resource):
    target = OUT / next(item["href"].split("/")[-1] for item in resource["files"] if item["format"] == "PDF")
    styles = pdf_styles()
    doc = SimpleDocTemplate(str(target), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=18 * mm, bottomMargin=19 * mm, title=resource["title"], author="Goway Danışmanlık", subject=resource["summary"])
    story = [
        Paragraph("GOWAY · UYGULAMA KAYNAĞI", styles["GSub"]),
        Paragraph(resource["title"], styles["GTitle"]),
        Paragraph(resource["summary"], styles["GBody"]),
        Spacer(1, 6 * mm),
    ]
    meta = Table([
        [Paragraph("Sürüm", styles["GSmall"]), Paragraph(resource["version"], styles["GSmall"]), Paragraph("Güncelleme", styles["GSmall"]), Paragraph(resource["updatedAt"], styles["GSmall"])],
        [Paragraph("Kapsam", styles["GSmall"]), Paragraph(resource["sector"], styles["GSmall"]), Paragraph("Teknik durum", styles["GSmall"]), Paragraph(resource["reviewer"], styles["GSmall"])],
    ], colWidths=[24 * mm, 52 * mm, 30 * mm, 60 * mm])
    meta.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#" + MIST)), ("GRID", (0, 0), (-1, -1), .35, colors.HexColor("#" + LIGHT)), ("FONTNAME", (0, 0), (-1, -1), FONT), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    story += [meta, Spacer(1, 7 * mm), Paragraph("Kullanım", styles["GHead"]), Paragraph("Her satırı evet, kısmen, hayır veya kapsam dışı olarak değerlendirin. Kanıt alanına belge adı yerine erişilebilir kayıt, tarih ve sorumlu bilgisi yazın. Açıkları sorumlu, termin ve doğrulama yöntemiyle aksiyona dönüştürün.", styles["GBody"]), Spacer(1, 5 * mm)]
    rows = [[Paragraph("Alan", styles["GSmall"]), Paragraph("Kontrol sorusu", styles["GSmall"]), Paragraph("Beklenen kanıt", styles["GSmall"]), Paragraph("Durum / not", styles["GSmall"])]]
    for area, question, evidence in RESOURCE_ROWS[resource["slug"]]:
        rows.append([Paragraph(area, styles["GSmall"]), Paragraph(question, styles["GSmall"]), Paragraph(evidence, styles["GSmall"]), Paragraph("□ Evet  □ Kısmen  □ Hayır<br/>Not: __________________", styles["GSmall"])])
    table = Table(rows, repeatRows=1, colWidths=[27 * mm, 65 * mm, 43 * mm, 40 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), .4, colors.HexColor("#B7CFDA")), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 7), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#" + MIST)])]))
    story += [table, Spacer(1, 6 * mm), Paragraph("Sınırlama", styles["GHead"]), Paragraph("Bu kaynak genel hazırlık ve öz değerlendirme amacıyla sunulur. İşletmeye özel mevzuat, risk değerlendirmesi, hukuki görüş, belgelendirme veya resmî uygunluk kararının yerine geçmez. Kullanıcı güncel gereklilikleri ve kendi saha koşullarını ayrıca doğrulamalıdır.", styles["GBody"])]
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


def base_workbook(title, instructions):
    wb = Workbook()
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws = wb.active
    ws.title = "Kullanım"
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=20, bold=True, color=NAVY)
    ws["A3"] = "Nasıl kullanılır?"
    ws["A3"].font = Font(name="Arial", size=12, bold=True, color=TEAL)
    ws["A4"] = instructions
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A6"] = "Sarı hücreler kullanıcı girdisidir. Formüllü hücreleri değiştirmeyin. Örnek satır, beklenen biçimi gösterir ve gerçek müşteri verisi değildir."
    ws["A6"].fill = PatternFill("solid", fgColor=INPUT)
    ws["A6"].alignment = Alignment(wrap_text=True)
    ws["A8"] = "Sürüm"
    ws["B8"] = "1.0"
    ws["A9"] = "Güncelleme"
    ws["B9"] = "2026-08-23"
    ws["A10"] = "Sahibi"
    ws["B10"] = "Goway Danışmanlık"
    ws["A11"] = "Teknik durum"
    ws["B11"] = "Teknik inceleme bekliyor"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 70
    for row in ws.iter_rows(min_row=1, max_row=11, min_col=1, max_col=2):
        for cell in row:
            cell.font = Font(name="Arial", size=10, bold=cell.column == 1, color=NAVY)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    return wb


def style_table(ws, widths):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[1].height = 34
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial", size=9, color="0000FF" if cell.data_type != "f" else "000000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if cell.data_type != "f":
                cell.fill = PatternFill("solid", fgColor=INPUT)


def make_planner(target):
    wb = base_workbook("30 Günlük Denetim Hazırlık Planı", "Plan sayfasındaki görevleri işletmenize göre düzenleyin. Sorumlu, termin, durum ve kanıt alanlarını güncel tutun. Özet sayfası formüllerle ilerlemeyi gösterir.")
    ws = wb.create_sheet("Plan")
    headers = ["Gün", "Hafta", "Görev", "Sorumlu", "Termin", "Durum", "Kanıt / bağlantı", "Not"]
    ws.append(headers)
    tasks = [row[1] for row in RESOURCE_ROWS["30-gunluk-denetim-hazirlik-plani"]]
    for day in range(1, 31):
        week = min(4, (day - 1) // 7 + 1)
        task = tasks[min(len(tasks) - 1, (day - 1) * len(tasks) // 30)]
        ws.append([day, f"{week}. hafta", task, "", "", "Başlanmadı", "", ""])
    validation = DataValidation(type="list", formula1='"Başlanmadı,Devam ediyor,Beklemede,Tamamlandı"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add("F2:F31")
    style_table(ws, [8, 12, 48, 22, 14, 18, 34, 32])
    for row in range(2, 32):
        ws.cell(row, 5).number_format = "dd.mm.yyyy"
    summary = wb.create_sheet("Özet")
    summary.append(["Gösterge", "Değer"])
    summary.append(["Toplam görev", "=COUNTA(Plan!A2:A31)"])
    summary.append(["Tamamlanan", '=COUNTIF(Plan!F2:F31,"Tamamlandı")'])
    summary.append(["İlerleme", "=IFERROR(B3/B2,0)"])
    summary["B4"].number_format = "0%"
    style_table(summary, [28, 18])
    wb.save(target)


def make_evidence_matrix(target):
    wb = base_workbook("Denetim Kanıt Matrisi", "Matris sayfasında gereklilikleri süreç, sorumlu ve gerçek kayıtlarla eşleştirin. Durumu seçin; özet formülleri açık kanıtları gösterir.")
    ws = wb.create_sheet("Matris")
    ws.append(["Madde / şart", "Gereklilik", "Süreç", "Sorumlu", "Kanıt / kayıt", "Revizyon / dönem", "Durum", "Aksiyon", "Termin"])
    examples = [
        ["ISO 9001 6.1", "Risk ve fırsatlar", "Stratejik planlama", "Kalite yöneticisi", "Risk kaydı 2026", "2026-Q3", "Kısmen", "Geciken aksiyonları kapat", ""],
        ["ISO 45001 8.1", "Operasyonel kontrol", "Bakım", "Bakım yöneticisi", "LOTO-PR-02 ve kartlar", "Rev.1", "Açık", "Ekipman kapsamını doğrula", ""],
        ["ISO 14001 9.1", "İzleme ve ölçme", "Çevre", "Çevre sorumlusu", "Atık ve emisyon kayıtları", "2026-Q2", "Tam", "", ""],
    ]
    for row in examples:
        ws.append(row)
    validation = DataValidation(type="list", formula1='"Açık,Kısmen,Tam,Kapsam dışı"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add("G2:G200")
    style_table(ws, [17, 34, 23, 22, 36, 18, 16, 34, 14])
    summary = wb.create_sheet("Özet")
    summary.append(["Gösterge", "Değer"])
    summary.append(["Dolu kayıt", "=COUNTA(Matris!A2:A200)"])
    summary.append(["Tam", '=COUNTIF(Matris!G2:G200,"Tam")'])
    summary.append(["Açık", '=COUNTIF(Matris!G2:G200,"Açık")'])
    summary.append(["Kısmen", '=COUNTIF(Matris!G2:G200,"Kısmen")'])
    summary.append(["Tamamlanma", "=IFERROR(B3/B2,0)"])
    summary["B6"].number_format = "0%"
    style_table(summary, [28, 18])
    wb.save(target)


def make_enpi(target):
    wb = base_workbook("ISO 50001 SEU ve EnPI Çalışma Kitabı", "Veri sayfasında sarı giriş hücrelerine dönem, enerji tüketimi, üretim/hizmet çıktısı, çalışma saati ve not girin. Enerji yoğunluğu formülle hesaplanır. Özet sayfası temel eğilimleri gösterir.")
    ws = wb.create_sheet("Veri")
    ws.append(["Dönem", "Enerji (kWh)", "Üretim / hizmet çıktısı", "Çalışma saati", "Enerji yoğunluğu (kWh/birim)", "SEU / proses", "Not"])
    examples = [
        ["2026-01", 128000, 42000, 510, "=IFERROR(B2/C2,0)", "Kompresör sistemi", "Örnek veri; gerçek sonuç değildir"],
        ["2026-02", 121500, 41500, 495, "=IFERROR(B3/C3,0)", "Kompresör sistemi", ""],
        ["2026-03", 119800, 43000, 505, "=IFERROR(B4/C4,0)", "Kompresör sistemi", ""],
    ]
    for row in examples:
        ws.append(row)
    for row in range(5, 62):
        ws.append(["", "", "", "", f"=IFERROR(B{row}/C{row},0)", "", ""])
    style_table(ws, [14, 18, 25, 18, 28, 24, 34])
    for row in range(2, 62):
        ws.cell(row, 5).number_format = "0.000"
        ws.cell(row, 5).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row, 5).font = Font(name="Arial", size=9, color="000000")
    summary = wb.create_sheet("Özet")
    summary.append(["Gösterge", "Değer"])
    summary.append(["Toplam enerji (kWh)", "=SUM(Veri!B2:B61)"])
    summary.append(["Toplam çıktı", "=SUM(Veri!C2:C61)"])
    summary.append(["Toplam yoğunluk", "=IFERROR(B2/B3,0)"])
    summary.append(["Ortalama dönem yoğunluğu", "=IFERROR(AVERAGE(Veri!E2:E61),0)"])
    summary.append(["En düşük dönem yoğunluğu", "=IFERROR(MIN(Veri!E2:E61),0)"])
    for cell in ["B4", "B5", "B6"]:
        summary[cell].number_format = "0.000"
    style_table(summary, [34, 20])
    wb.save(target)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    for resource in RESOURCES:
        make_pdf(resource)
    make_planner(OUT / "30-gunluk-denetim-hazirlik-plani-v1.xlsx")
    make_evidence_matrix(OUT / "denetim-kanit-matrisi-v1.xlsx")
    make_enpi(OUT / "iso-50001-seu-enpi-calisma-kitabi-v1.xlsx")
    print(json.dumps({"pdf": len(RESOURCES), "xlsx": 3, "output": str(OUT)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
