from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent.parent
sys.stdout.reconfigure(encoding="utf-8")
OUT = ROOT / "assets" / "downloads"
manifest = json.loads((ROOT / "data" / "resources.json").read_text(encoding="utf-8"))["resources"]
expected = {Path(item["href"]).name for resource in manifest for item in resource["files"]}
failures = []
results = {"pdf": [], "docx": [], "xlsx": []}

for name in sorted(expected):
    file = OUT / name
    if not file.exists():
        failures.append(f"missing: {name}")
        continue
    if file.stat().st_size < 5000:
        failures.append(f"unexpectedly small: {name}")

for file in sorted(OUT.glob("*.pdf")):
    reader = PdfReader(file)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not reader.pages or len(text.strip()) < 300:
        failures.append(f"PDF content missing: {file.name}")
    if "Sınırlama" not in text:
        failures.append(f"PDF disclaimer missing: {file.name}")
    results["pdf"].append({"file": file.name, "pages": len(reader.pages), "characters": len(text)})

for file in sorted(OUT.glob("*.docx")):
    try:
        with zipfile.ZipFile(file) as archive:
            names = set(archive.namelist())
            required = {"[Content_Types].xml", "word/document.xml", "word/styles.xml"}
            if not required.issubset(names):
                failures.append(f"DOCX package incomplete: {file.name}")
            document = archive.read("word/document.xml")
            ElementTree.fromstring(document)
            plain = b" ".join(re.findall(rb"<w:t[^>]*>(.*?)</w:t>", document)).decode("utf-8", errors="replace")
            if len(plain) < 250 or "Sınırlama" not in plain:
                failures.append(f"DOCX content incomplete: {file.name}")
            results["docx"].append({"file": file.name, "characters": len(plain), "parts": len(names)})
    except Exception as exc:
        failures.append(f"DOCX parse failed {file.name}: {exc}")

banned = ("XLOOKUP", "XMATCH", "FILTER(", "UNIQUE(", "SEQUENCE(")
for file in sorted(OUT.glob("*.xlsx")):
    workbook = load_workbook(file, data_only=False)
    formulas = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(f"{sheet.title}!{cell.coordinate}:{cell.value}")
    if not formulas:
        failures.append(f"XLSX formulas missing: {file.name}")
    if any(any(token in formula.upper() for token in banned) for formula in formulas):
        failures.append(f"XLSX unsupported formula: {file.name}")
    if workbook.calculation.calcMode != "auto" or not workbook.calculation.fullCalcOnLoad:
        failures.append(f"XLSX recalculation flags missing: {file.name}")
    results["xlsx"].append({"file": file.name, "sheets": workbook.sheetnames, "formulas": len(formulas), "calcMode": workbook.calculation.calcMode})

unexpected = {file.name for file in OUT.iterdir() if file.is_file()} - expected
if unexpected:
    failures.append(f"unexpected files: {sorted(unexpected)}")

print(json.dumps({"ok": not failures, "failures": failures, "results": results}, ensure_ascii=False, indent=2))
raise SystemExit(1 if failures else 0)
